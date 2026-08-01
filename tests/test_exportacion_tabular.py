"""Paso 6: las exportaciones tabulares (CSV y XLSX).

Qué se fija aquí, y por qué cada cosa.

El CSV se probaba antes con una sola aserción sobre su cabecera, así que nada
impedía que un campo JSON acabara volcado en una celda ni que un nombre de
perfil ajeno se convirtiera en una fórmula. Este fichero cubre los dos formatos
por el mismo sitio, porque **comparten columnas por construcción**
(`tabular.COLUMNAS`) y esa igualdad es justo lo que hay que vigilar: el día que
alguien añada una columna a uno solo, el test de paridad lo dice.

Tres bloques que no son opcionales:

  - **Inyección de fórmulas.** `display_name` y el nombre del sitio vienen de
    páginas ajenas. Si una celda empieza por `=`, `+`, `-` o `@`, Excel y
    LibreOffice la ejecutan al abrir el fichero. Hay un test de barrido que
    recorre TODAS las columnas de una cuenta envenenada y falla si alguna se
    escapa sin neutralizar — no basta con probar el caso que se nos ocurrió.
  - **Límites de Excel.** Medido: openpyxl guarda una celda de 40.000
    caracteres sin avisar y es Excel quien luego rechaza el fichero. Y un
    título de hoja largo solo produce un `UserWarning`. Los dos se sanean
    aquí, y se comprueba que el recorte SE VE.
  - **Transporte HTTP.** La lección del Paso 5: se probaba el generador y nunca
    el transporte, y el bug vivía justo ahí. `format=xlsx` se pide por HTTP.
"""
from __future__ import annotations

import codecs
import csv
import io
import json
import re
import time
from pathlib import Path

from .helpers import IsolatedTestCase, auth_client

_RAIZ = Path(__file__).resolve().parents[1]
_INDEX_HTML = _RAIZ / "rastrillo" / "static" / "index.html"

# Los cuatro prefijos peligrosos, y las variantes que llevan delante algo que
# los lectores saltan antes de decidir si la celda es una fórmula.
PELIGROSOS = ("=1+1", "+1+1", "-1+1", "@SUM(A1)",
              "\t=cmd|'/c calc'!A1", "\r=1+1", " =1+1")


class _Base(IsolatedTestCase):
    def setUp(self):
        super().setUp()
        from rastrillo import db, reports, tabular
        self.db = db
        self.reports = reports
        self.tabular = tabular
        self.db.init()

    def alta(self, **campos):
        base = dict(
            source="sherlock", source_site="ejemplo.com",
            profile_url="https://ejemplo.com/u/mar", display_name="Ejemplo",
            confidence="high", status="found",
        )
        base.update(campos)
        plataforma = base.pop("platform", "ejemplo")
        ident = base.pop("identifier", "mar")
        return self.db.upsert_account(plataforma, ident, **base)

    # -- utilidades ---------------------------------------------------------
    def csv_bytes(self, sep=None) -> bytes:
        contenido, _mt, _n = self.reports.build_report("csv", sep)
        return contenido

    def csv_filas(self, sep=","):
        """El CSV parseado de vuelta con el `csv` de stdlib."""
        texto = self.csv_bytes(sep).decode("utf-8-sig")
        return list(csv.reader(io.StringIO(texto, newline=""), delimiter=sep))

    def libro(self):
        from openpyxl import load_workbook
        contenido, _mt, _n = self.reports.build_report("xlsx")
        return load_workbook(io.BytesIO(contenido))

    def celdas_texto(self, hoja):
        """Todos los valores de texto de una hoja, sin la cabecera."""
        return [c.value for fila in hoja.iter_rows(min_row=2) for c in fila
                if isinstance(c.value, str)]


# ── CSV (tests 1-6) ─────────────────────────────────────────────────────────
class TestCSVComoCSV(_Base):

    def test_1_bom_y_crlf(self):
        """BOM UTF-8 al inicio y terminaciones \\r\\n, como manda el estándar.

        Sin BOM, Excel lee el fichero en la codificación ANSI del sistema y
        destroza acentos y cirílico.
        """
        self.alta()
        datos = self.csv_bytes()
        self.assertTrue(datos.startswith(codecs.BOM_UTF8))
        cuerpo = datos[len(codecs.BOM_UTF8):]
        self.assertIn(b"\r\n", cuerpo)
        # Ningún \n suelto: todos tienen que venir precedidos de \r.
        self.assertEqual(cuerpo.count(b"\n"), cuerpo.count(b"\r\n"))

    def test_2_round_trip_con_stdlib(self):
        self.alta(identifier="mar", display_name="Ejemplo")
        self.alta(platform="otro", identifier="je7remy",
                  source_site="otro.com", display_name="Otro")
        filas = self.csv_filas()
        self.assertEqual(filas[0], list(self.tabular.TITULOS))
        self.assertEqual(len(filas), 3, "cabecera + 2 cuentas")
        col = filas[0].index("Identificador")
        self.assertEqual({filas[1][col], filas[2][col]}, {"mar", "je7remy"})

    def test_3_separador_comillas_y_salto_sobreviven(self):
        """Un valor con el separador, comillas y un salto → se escapa bien."""
        veneno = 'nota con , coma; "comillas" y\nun salto'
        self.alta(last_message=veneno)
        filas = self.csv_filas()
        col = filas[0].index("Última nota")
        self.assertEqual(filas[1][col], veneno)

    def test_3b_el_mismo_valor_con_separador_punto_y_coma(self):
        """Y con el otro separador, que es el que usará quien tenga Excel es-ES."""
        veneno = 'a;b,c "d" e\nf'
        self.alta(last_message=veneno)
        filas = self.csv_filas(sep=";")
        col = filas[0].index("Última nota")
        self.assertEqual(filas[1][col], veneno)

    def test_4_cirilico_y_acentos(self):
        self.alta(display_name="Малыши и мамы", identifier="je7remy",
                  last_message="anonimización con tildes: ñ á é ü")
        filas = self.csv_filas()
        self.assertIn("Малыши и мамы", filas[1])
        self.assertIn("anonimización con tildes: ñ á é ü", filas[1])

    def test_5_ningun_campo_json_crudo(self):
        """Ni `confidence_reasons`, ni `action_meta`, ni `breach_meta`.

        Un `{"code": "bump_path"}` dentro de una celda es ilegible y además
        rompe el parseo de quien lea el fichero.
        """
        self.alta(
            confidence_reasons=self.db.dump_reasons(
                [{"code": "bump_path", "desc": "coincide en la ruta"},
                 {"code": "tramo_distintivo", "desc": "username distintivo"}]),
            action_meta=json.dumps({"kind": "semi_auto", "layer": "directory",
                                    "url": "https://ejemplo.com/baja"}),
            breach_meta=json.dumps({"breach_date": "2019-03-01",
                                    "pwn_count": 1234567,
                                    "data_classes": ["Passwords", "Genders"],
                                    "is_verified": True}),
        )
        filas = self.csv_filas()
        for celda in filas[1]:
            self.assertNotIn('{"', celda, f"JSON crudo en una celda: {celda!r}")
            self.assertNotIn('"code"', celda)
            self.assertNotIn("data_classes", celda)
        # Y lo que sí tiene que haber: los motivos como texto legible.
        senales = filas[1][filas[0].index("Señales")]
        self.assertIn("coincide en la ruta", senales)
        self.assertIn("username distintivo", senales)

    def test_5b_breach_meta_sale_como_columnas_propias(self):
        self.alta(breach_meta=json.dumps({
            "breach_date": "2019-03-01", "pwn_count": 1234567,
            "data_classes": ["Passwords", "Email addresses"],
            "is_verified": False, "is_spam_list": True}))
        filas = self.csv_filas()
        fila = dict(zip(filas[0], filas[1]))
        self.assertEqual(fila["Brecha: fecha"], "2019-03-01")
        self.assertEqual(fila["Brecha: cuentas expuestas"], "1234567")
        self.assertIn("Contraseñas", fila["Brecha: datos expuestos"])
        self.assertIn("lista de spam", fila["Brecha: advertencias"])
        self.assertIn("sin verificar", fila["Brecha: advertencias"])

    def test_6_db_vacia_solo_cabecera(self):
        filas = self.csv_filas()
        self.assertEqual(len(filas), 1)
        self.assertEqual(filas[0], list(self.tabular.TITULOS))

    def test_perfil_y_baja_son_columnas_distintas(self):
        """El error que ya se arregló en la UI y en el PDF no vuelve por aquí."""
        self.alta(profile_url="https://ejemplo.com/u/mar",
                  action_meta=json.dumps({"url": "https://ejemplo.com/baja"}))
        filas = self.csv_filas()
        fila = dict(zip(filas[0], filas[1]))
        self.assertEqual(fila["Perfil detectado"], "https://ejemplo.com/u/mar")
        self.assertEqual(fila["Cómo darse de baja"], "https://ejemplo.com/baja")

    def test_ninguna_celda_dice_None(self):
        self.alta(display_name=None, last_message=None, profile_url=None)
        filas = self.csv_filas()
        for celda in filas[1]:
            self.assertNotEqual(celda, "None")
            self.assertNotEqual(celda, "none")

    def test_estados_traducidos_no_crudos(self):
        self.alta(status="not_mine")
        filas = self.csv_filas()
        fila = dict(zip(filas[0], filas[1]))
        self.assertEqual(fila["Estado"], "Descartada")
        self.assertNotIn("not_mine", filas[1])

    def test_fechas_legibles_no_timestamps(self):
        self.alta(sent_at=1785554447.29)
        filas = self.csv_filas()
        fila = dict(zip(filas[0], filas[1]))
        self.assertRegex(fila["Solicitud enviada el"],
                         r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}$")
        self.assertNotIn("1785554447", filas[1])

    def test_separador_por_defecto_es_la_coma(self):
        """RFC 4180. El knob existe, pero no cambia el default."""
        self.alta()
        cabecera = self.csv_bytes().decode("utf-8-sig").splitlines()[0]
        self.assertIn(",", cabecera)
        self.assertNotIn(";", cabecera)

    def test_no_hay_linea_sep_de_microsoft(self):
        """`sep=;` añadiría una fila espuria a cualquier parser estándar."""
        self.alta()
        primera = self.csv_bytes().decode("utf-8-sig").splitlines()[0]
        self.assertFalse(primera.lower().startswith("sep="))


# ── XLSX (tests 7-12) ───────────────────────────────────────────────────────
class TestXLSX(_Base):

    def test_7_abre_y_tiene_las_hojas(self):
        self.alta()
        wb = self.libro()
        self.assertEqual(wb.sheetnames, ["Resumen", "Cuentas", "Glosario"])

    def test_7b_empieza_por_PK(self):
        """Un xlsx es un zip."""
        contenido, media, nombre = self.reports.build_report("xlsx")
        self.assertTrue(contenido.startswith(b"PK"))
        self.assertIn("spreadsheetml", media)
        self.assertTrue(nombre.endswith(".xlsx"))

    def test_8_fechas_son_fechas_y_numeros_son_numeros(self):
        import datetime as dt
        self.alta(sent_at=1785554447.29,
                  breach_meta=json.dumps({"breach_date": "2019-03-01",
                                          "pwn_count": 1234567}))
        hoja = self.libro()["Cuentas"]
        titulos = [c.value for c in hoja[1]]
        fila = {t: hoja.cell(row=2, column=i + 1)
                for i, t in enumerate(titulos)}

        self.assertIsInstance(fila["Solicitud enviada el"].value, dt.datetime)
        self.assertIsInstance(fila["Brecha: fecha"].value, (dt.date, dt.datetime))
        # `pwn_count` tiene que ser sumable en Excel, no una cadena.
        self.assertIsInstance(fila["Brecha: cuentas expuestas"].value, int)
        self.assertEqual(fila["Brecha: cuentas expuestas"].value, 1234567)
        self.assertIn("#,##0", fila["Brecha: cuentas expuestas"].number_format)
        self.assertIsInstance(fila["id"].value, int)

    def test_9_cabecera_congelada_y_autofiltro(self):
        self.alta()
        hoja = self.libro()["Cuentas"]
        self.assertEqual(hoja.freeze_panes, "A2")
        self.assertIsNotNone(hoja.auto_filter.ref)
        self.assertTrue(hoja.auto_filter.ref.startswith("A1:"))

    def test_9b_cabecera_se_distingue_por_peso_no_por_color(self):
        self.alta()
        hoja = self.libro()["Cuentas"]
        self.assertTrue(all(c.font.bold for c in hoja[1]))

    def test_10_db_vacia_da_fichero_valido(self):
        wb = self.libro()
        self.assertEqual(wb.sheetnames, ["Resumen", "Cuentas", "Glosario"])
        hoja = wb["Cuentas"]
        self.assertEqual([c.value for c in hoja[1]],
                         list(self.tabular.TITULOS))
        self.assertIsNotNone(hoja.auto_filter.ref)

    def test_11_300_filas_en_tiempo_razonable(self):
        for i in range(300):
            self.alta(platform=f"sitio{i}", identifier=f"usuario{i}",
                      source_site=f"sitio{i}.com",
                      display_name=f"Sitio número {i}")
        t0 = time.time()
        contenido, _mt, _n = self.reports.build_report("xlsx")
        tardanza = time.time() - t0
        self.assertLess(tardanza, 30, f"el XLSX tardó {tardanza:.1f}s")
        from openpyxl import load_workbook
        hoja = load_workbook(io.BytesIO(contenido))["Cuentas"]
        self.assertEqual(hoja.max_row, 301)

    def test_12_nulos_y_cadenas_patologicas(self):
        self.alta(display_name=None, profile_url=None, last_message=None,
                  confidence=None, status="found")
        self.alta(platform="raro", identifier="Малыши",
                  display_name="Ру́сский 中文 🎉",
                  profile_url="https://ejemplo.com/" + "x" * 600,
                  last_message="línea\ncon salto\ty tab")
        self.alta(platform="basura", identifier="z",
                  confidence_reasons="{esto no es JSON",
                  breach_meta="tampoco[",
                  action_meta="{roto")
        wb = self.libro()          # no lanza
        hoja = wb["Cuentas"]
        self.assertEqual(hoja.max_row, 4)
        for celda in self.celdas_texto(hoja):
            self.assertNotEqual(celda, "None")

    def test_el_alcance_no_enumera_300_identificadores(self):
        """La misma regresión que el `MAX_ALCANCE` del PDF.

        Allí reventaba el layout con un `LayoutError`; aquí "solo" deja una
        celda de miles de caracteres que nadie puede leer. El remedio es el
        mismo y hay que fijarlo igual, porque con nombres cortos no se nota.
        """
        for i in range(300):
            self.alta(platform=f"p{i}", identifier=f"identificador-larguito-{i}",
                      source_site=f"s{i}.com")
        hoja = self.libro()["Resumen"]
        textos = [c.value for fila in hoja.iter_rows() for c in fila
                  if isinstance(c.value, str)]
        alcance = [t for t in textos if "identificador-larguito-" in t]
        self.assertTrue(alcance, "el resumen no enseña el alcance")
        for t in alcance:
            self.assertLess(len(t), 800, "el alcance enumera demasiados")
            self.assertIn(" más", t, "debería decir cuántos quedan fuera")

    def test_hojas_de_resumen_y_glosario_tienen_contenido(self):
        self.alta()
        wb = self.libro()
        for nombre in ("Resumen", "Glosario"):
            textos = [c.value for fila in wb[nombre].iter_rows() for c in fila
                      if isinstance(c.value, str)]
            self.assertTrue(textos, f"la hoja {nombre} está vacía")

    def test_el_glosario_sale_de_app_js(self):
        """Los mismos textos que el anexo del PDF y que los tooltips."""
        from rastrillo import glosario
        self.alta()
        textos = [c.value for fila in self.libro()["Glosario"].iter_rows()
                  for c in fila if isinstance(c.value, str)]
        esperado = glosario.explicacion_confianza("low")
        self.assertTrue(esperado, "app.js no trae el tooltip de `low`")
        self.assertIn(esperado, textos)
        # Y la frase que importa: `low` no dice "no es tuya".
        self.assertIn("DÉBIL", esperado)

    def test_urls_son_hipervinculos(self):
        self.alta(profile_url="https://ejemplo.com/u/mar")
        hoja = self.libro()["Cuentas"]
        titulos = [c.value for c in hoja[1]]
        celda = hoja.cell(row=2, column=titulos.index("Perfil detectado") + 1)
        self.assertEqual(celda.hyperlink.target, "https://ejemplo.com/u/mar")

    def test_anchos_de_columna_no_son_todos_iguales(self):
        self.alta(last_message="una nota bastante más larga que el resto "
                               "de los campos de esta fila, para que el "
                               "ancho medido salga distinto")
        hoja = self.libro()["Cuentas"]
        anchos = {d.width for d in hoja.column_dimensions.values()}
        self.assertGreater(len(anchos), 1, "todas las columnas miden igual")


# ── Paridad CSV / XLSX ──────────────────────────────────────────────────────
class TestParidad(_Base):
    """Los dos formatos enseñan lo mismo. Es el invariante de `tabular.py`."""

    def test_mismas_columnas_en_el_mismo_orden(self):
        self.alta()
        cabecera_csv = self.csv_filas()[0]
        cabecera_xlsx = [c.value for c in self.libro()["Cuentas"][1]]
        self.assertEqual(cabecera_csv, cabecera_xlsx)

    def test_mismos_valores_de_texto(self):
        self.alta(identifier="je7remy", display_name="Дуолинго",
                  status="email_draft", last_message="nota con , coma")
        fila_csv = dict(zip(*self.csv_filas()[:2]))
        hoja = self.libro()["Cuentas"]
        titulos = [c.value for c in hoja[1]]
        for i, titulo in enumerate(titulos):
            valor = hoja.cell(row=2, column=i + 1).value
            if isinstance(valor, str):
                self.assertEqual(fila_csv[titulo], valor,
                                 f"la columna {titulo!r} difiere entre formatos")


# ── Inyección de fórmulas y límites (tests 13-15) ───────────────────────────
class TestInyeccionYLimites(_Base):

    def test_13_csv_neutraliza_los_prefijos(self):
        for veneno in PELIGROSOS:
            with self.subTest(veneno=veneno):
                self.db.clear_accounts()
                self.alta(display_name=veneno)
                filas = self.csv_filas()
                celda = filas[1][filas[0].index("Sitio")]
                self.assertTrue(
                    celda.startswith("'"),
                    f"{veneno!r} salió sin neutralizar como {celda!r}")
                # Round-trip: el dato sigue siendo reconocible.
                self.assertEqual(celda.lstrip("'"),
                                 self.tabular.limpiar_control(veneno).strip())

    def test_13b_xlsx_los_escribe_como_texto(self):
        """En XLSX no hace falta alterar el dato: basta el tipo de celda."""
        for veneno in PELIGROSOS:
            with self.subTest(veneno=veneno):
                self.db.clear_accounts()
                self.alta(display_name=veneno)
                hoja = self.libro()["Cuentas"]
                titulos = [c.value for c in hoja[1]]
                celda = hoja.cell(row=2, column=titulos.index("Sitio") + 1)
                self.assertEqual(celda.data_type, "s",
                                 f"{veneno!r} quedó como fórmula en el XLSX")
                self.assertNotIn(celda.value[:1], ("",),)

    def test_13c_barrido_ninguna_columna_se_escapa(self):
        """El test que importa: se envenenan TODAS las columnas de texto.

        Probar solo el campo que se nos ocurrió deja la puerta abierta a la
        siguiente columna que alguien añada. Aquí falla si CUALQUIERA de ellas
        deja pasar un valor sin neutralizar, en cualquiera de los dos formatos.
        """
        veneno = "=HYPERLINK(\"http://malo\",\"pincha\")"
        self.alta(
            display_name=veneno, source_site=veneno, identifier=veneno,
            source=veneno, last_message=veneno, deletion_type=veneno,
            difficulty=veneno, profile_url=veneno,
            action_meta=json.dumps({"url": veneno, "layer": veneno,
                                    "kind": veneno, "email_to": veneno,
                                    "email_subject": veneno}),
            breach_meta=json.dumps({"data_classes": [veneno]}),
        )
        filas = self.csv_filas()
        for titulo, celda in zip(filas[0], filas[1]):
            if self.tabular.empieza_por_formula(celda):
                self.fail(f"CSV: la columna {titulo!r} dejó pasar {celda!r}")

        hoja = self.libro()["Cuentas"]
        titulos = [c.value for c in hoja[1]]
        for i, titulo in enumerate(titulos):
            celda = hoja.cell(row=2, column=i + 1)
            if isinstance(celda.value, str) and \
                    self.tabular.empieza_por_formula(celda.value):
                self.assertEqual(
                    celda.data_type, "s",
                    f"XLSX: la columna {titulo!r} quedó como fórmula")

    def test_13d_un_valor_inocente_no_se_toca(self):
        """La guarda no puede ensuciar lo que no es peligroso."""
        self.alta(display_name="Duolingo", last_message="todo normal")
        filas = self.csv_filas()
        fila = dict(zip(filas[0], filas[1]))
        self.assertEqual(fila["Sitio"], "Duolingo")
        self.assertEqual(fila["Última nota"], "todo normal")

    def test_14_cadena_de_40000_se_recorta_visiblemente(self):
        largo = "L" * 40_000
        self.alta(last_message=largo)

        filas = self.csv_filas()
        celda = filas[1][filas[0].index("Última nota")]
        self.assertLessEqual(len(celda), self.tabular.MAX_CELDA)
        self.assertIn("recortado", celda)
        self.assertIn("caracteres omitidos", celda)

        hoja = self.libro()["Cuentas"]
        titulos = [c.value for c in hoja[1]]
        valor = hoja.cell(row=2, column=titulos.index("Última nota") + 1).value
        self.assertLessEqual(len(valor), 32_767, "Excel rechazaría la celda")
        self.assertIn("recortado", valor)

    def test_14b_recortar_respeta_el_limite_exacto(self):
        for n in (0, 1, 50, 999, 32_000, 40_000):
            with self.subTest(n=n):
                salida = self.tabular.recortar("x" * n, limite=100)
                self.assertLessEqual(len(salida), 100)
                if n > 100:
                    self.assertIn("omitidos", salida)
                else:
                    self.assertEqual(salida, "x" * n)

    def test_14c_recortar_cuenta_las_incidencias(self):
        inc = self.tabular.Incidencias()
        self.tabular.recortar("x" * 200, limite=50, incidencias=inc)
        self.tabular.recortar("corto", limite=50, incidencias=inc)
        self.assertEqual(inc.recortados, 1)

    def test_15_nombre_de_hoja_saneado(self):
        for bruto, esperado in (
            ("Cuentas", "Cuentas"),
            ("a[b]c", "abc"),
            ("a/b\\c:d*e?f", "abcdef"),
            ("", "Hoja"),
            (None, "Hoja"),
            ("con\x00control", "concontrol"),
        ):
            with self.subTest(bruto=bruto):
                self.assertEqual(self.tabular.nombre_hoja(bruto), esperado)

    def test_15b_nombre_de_hoja_largo_se_corta_a_31(self):
        largo = "Sitio con un nombre larguísimo que no cabe"
        salida = self.tabular.nombre_hoja(largo)
        self.assertLessEqual(len(salida), 31)
        self.assertTrue(salida.endswith("…"), "el corte tiene que verse")

    def test_15c_openpyxl_acepta_lo_que_saneamos(self):
        """Lo que devuelve `nombre_hoja` nunca puede hacer estallar a openpyxl."""
        from openpyxl import Workbook
        wb = Workbook()
        for bruto in ("a[b]", "x" * 60, "a/b", "", "?:*"):
            with self.subTest(bruto=bruto):
                wb.create_sheet(self.tabular.nombre_hoja(bruto, "Hoja1")
                                + str(len(wb.sheetnames)))

    def test_control_chars_no_tumban_el_xlsx(self):
        """openpyxl lanza IllegalCharacterError; se limpian antes."""
        self.alta(display_name="malo\x00\x07aquí", last_message="\x0bvertical")
        self.libro()       # no lanza
        filas = self.csv_filas()
        self.assertIn("maloaquí", filas[1])


# ── HTTP (tests 16-18) ──────────────────────────────────────────────────────
class TestHTTP(_Base):
    def setUp(self):
        super().setUp()
        self.cli = auth_client()

    def test_16_xlsx_por_http(self):
        self.alta()
        r = self.cli.get("/api/report?format=xlsx", headers=self.hdr())
        self.assertEqual(r.status_code, 200)
        self.assertEqual(
            r.headers["content-type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertTrue(r.content.startswith(b"PK"),
                        f"el cuerpo no es un xlsx: {r.content[:20]!r}")
        cd = r.headers.get("content-disposition", "")
        self.assertIn("attachment", cd)
        self.assertTrue(re.search(r'filename="rastrillo-[^"]+\.xlsx"', cd), cd)

    def test_16b_xlsx_nunca_devuelve_json(self):
        self.alta()
        r = self.cli.get("/api/report?format=xlsx", headers=self.hdr())
        self.assertNotIn("json", r.headers["content-type"].lower())
        self.assertFalse(r.content.lstrip().startswith(b"{"))

    def test_17_sin_token_401_y_sin_fichero(self):
        self.alta(identifier="secreto-mio", source_site="privado.example")
        r = self.cli.get("/api/report?format=xlsx")
        self.assertEqual(r.status_code, 401)
        self.assertFalse(r.content.startswith(b"PK"))
        self.assertNotIn(b"secreto-mio", r.content)

    def test_18_el_html_sigue_sin_anclas_al_informe(self):
        html = _INDEX_HTML.read_text(encoding="utf-8")
        anclas = re.findall(r"<a\b[^>]*>", html, flags=re.I | re.S)
        self.assertEqual([a for a in anclas if "/api/report" in a], [])

    def test_18b_el_boton_de_excel_esta_enganchado(self):
        html = _INDEX_HTML.read_text(encoding="utf-8")
        self.assertIn("report-xlsx-btn", html)
        js = (_RAIZ / "rastrillo" / "static" / "app.js").read_text(
            encoding="utf-8")
        self.assertRegex(
            js, r'\$\("report-xlsx-btn"\)\.onclick\s*=.*descargarInforme\("xlsx"')

    def test_formato_desconocido_sigue_dando_400(self):
        for fmt in ("xml", "ods", "xls", "yaml"):
            with self.subTest(fmt=fmt):
                r = self.cli.get(f"/api/report?format={fmt}",
                                 headers=self.hdr())
                self.assertEqual(r.status_code, 400)

    def test_sep_por_query_cambia_el_separador_del_csv(self):
        self.alta()
        r = self.cli.get("/api/report?format=csv&sep=%3B", headers=self.hdr())
        self.assertEqual(r.status_code, 200)
        cabecera = r.content.decode("utf-8-sig").splitlines()[0]
        self.assertIn(";", cabecera)
        self.assertNotIn(",", cabecera)     # ningún título lleva coma

    def test_los_otros_formatos_no_cambian(self):
        """Los contratos de json y pdf quedan intactos."""
        self.alta()
        r = self.cli.get("/api/report?format=json", headers=self.hdr())
        self.assertIn("application/json", r.headers["content-type"])
        datos = r.json()
        self.assertIn("accounts", datos)
        self.assertIn("summary", datos)
        r = self.cli.get("/api/report?format=pdf", headers=self.hdr())
        self.assertTrue(r.content.startswith(b"%PDF"))
