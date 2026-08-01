"""El informe XLSX (openpyxl).

Lo usa el endpoint `GET /api/report?format=xlsx` y el subcomando
`rastrillo report --format xlsx --out FILE`. Firma paralela a
`report_pdf.render_pdf` para que `reports.build_report` los trate igual.

Por qué existe un cuarto formato
--------------------------------
El CSV no admite formato: no tiene tipografía, ni anchos de columna, ni tipos
de celda. Es texto plano cuya función es que otro programa lo lea, y meterle
adornos lo rompería como formato de intercambio, que es justo su valor. Así que
el fichero "para mirar" tenía que ser otro. Este.

Reparto: **CSV para procesar, XLSX para mirar.** Las columnas son las MISMAS
—las dos salen de `tabular.COLUMNAS`— y hay un test que lo exige. Lo que cambia
es que aquí una fecha es una fecha, un número es sumable, la cabecera se queda
fija al hacer scroll y hay filtros.

Criterios de diseño (los mismos que el PDF, y por las mismas razones)
---------------------------------------------------------------------
- **Ningún dato depende del color.** Estado, confianza y verificabilidad se
  escriben SIEMPRE como palabras, y cada recuento del resumen lleva su número.
  Si alguien imprime la hoja en blanco y negro, se lee igual. La cabecera se
  distingue por PESO (negrita + un filete), no por un bloque de color saturado.
- **La URL del PERFIL y la de BAJA son columnas separadas y etiquetadas.** Es
  el mismo error que se arregló en la UI y en el PDF, y no vuelve por aquí.
- **Nada de `None` impreso.** Un campo vacío es una celda vacía.
- **Lo que se altera, se declara.** La hoja de resumen dice cuántos valores se
  recortaron y cuántos se neutralizaron. Misma política que `pdf_fuentes`.

Las tres hojas
--------------
1. **Resumen** — totales, distribuciones y las notas del informe. Es la
   portada + el resumen del PDF.
2. **Cuentas** — una fila por cuenta, las columnas del CSV. Las pestañas del
   dashboard ("Descartadas", "Brechas"…) son filtros sobre esta misma tabla, y
   el autofiltro las reproduce sin duplicar filas — por eso no hay una hoja por
   estado ni una hoja de brechas aparte.
3. **Glosario** — las mismas explicaciones que el anexo del PDF, leídas de
   `static/app.js` vía `glosario.py`. Si no se pueden leer, la hoja lo dice en
   vez de inventárselas.

Lo que openpyxl NO protege (medido en la 3.1.5, por eso lo hacemos nosotros)
---------------------------------------------------------------------------
- Una celda de más de 32.767 caracteres **se guarda sin avisar**; es Excel quien
  luego rechaza el fichero. `tabular.recortar` corta antes y deja dicho cuánto.
- Un título de hoja de más de 31 caracteres solo emite un `UserWarning` y se
  conserva. `tabular.nombre_hoja` lo corta.
- Una cadena que empieza por `=`, `+`, `-` o `@` se convierte en FÓRMULA sola
  (`data_type` pasa a `"f"`). Se fuerza a `"s"`; ver `_escribir`.
"""
from __future__ import annotations

import io
import time as _t
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import glosario, tabular

# ── Presentación ────────────────────────────────────────────────────────────
# Un gris muy claro para la cabecera: suficiente para separarla de los datos,
# lejos de un bloque saturado que en una impresión sale como una mancha.
RELLENO_CABECERA = PatternFill("solid", fgColor="FFEFEFEF")
FILETE = Side(style="thin", color="FFBBBBBB")

FUENTE_CABECERA = Font(bold=True)
FUENTE_TITULO = Font(bold=True, size=14)
FUENTE_GRUPO = Font(bold=True, size=11)
FUENTE_NOTA = Font(italic=True, color="FF666666")

FORMATO_FECHA_HORA = "yyyy-mm-dd hh:mm"
FORMATO_FECHA = "yyyy-mm-dd"
FORMATO_ENTERO = "#,##0"

# Anchos de columna: se miden sobre el contenido real, con suelo y techo. Sin
# techo, una URL de 500 caracteres deja una columna que no cabe en la pantalla.
ANCHO_MIN = 10
ANCHO_MAX = 52

# Hipervínculo solo si es una URL de verdad. Un valor raro en la columna de URL
# se queda como texto: mejor eso que un enlace roto que Excel marque.
_ESQUEMAS = ("http://", "https://", "mailto:")

# Identificadores que se enumeran en el «Alcance» del resumen. Es el mismo tope
# que `report_pdf.MAX_ALCANCE` y por la misma razón: con 300 cuentas la lista
# completa es una celda de miles de caracteres que nadie lee. En el PDF además
# reventaba el layout; aquí "solo" es ilegible, pero el remedio es el mismo.
MAX_ALCANCE = 12


def _lista_resumida(items: Sequence[str], maximo: int = MAX_ALCANCE) -> str:
    """`a, b, c y 297 más`. Enumerar 300 no informa."""
    items = list(items)
    if len(items) <= maximo:
        return ", ".join(items)
    return f"{', '.join(items[:maximo])} y {len(items) - maximo} más"


def _escribir(celda, valor: Any) -> None:
    """Escribe un valor en una celda respetando su tipo, sin fórmulas.

    El orden importa: openpyxl decide `data_type` al ASIGNAR el valor, así que
    la corrección tiene que ir después. Con `data_type = "s"` el valor queda
    byte-idéntico y Excel lo muestra como texto — es decir, el dato no se
    altera en absoluto, que es mejor de lo que se puede hacer en un CSV.
    """
    if valor is None or valor == "":
        return
    celda.value = valor
    if isinstance(valor, datetime):
        celda.number_format = FORMATO_FECHA_HORA
    elif isinstance(valor, date):
        celda.number_format = FORMATO_FECHA
    elif isinstance(valor, bool):
        pass                       # se escribe como booleano; no es fórmula
    elif isinstance(valor, int):
        celda.number_format = FORMATO_ENTERO
    elif tabular.empieza_por_formula(valor):
        celda.data_type = "s"


def _fila_cabecera(hoja, titulos: Sequence[str], fila: int = 1) -> None:
    """Cabecera: negrita, filete y texto arriba. Sin color saturado."""
    for i, titulo in enumerate(titulos, start=1):
        c = hoja.cell(row=fila, column=i, value=titulo)
        c.font = FUENTE_CABECERA
        c.fill = RELLENO_CABECERA
        c.border = Border(bottom=FILETE)
        c.alignment = Alignment(vertical="top", wrap_text=False)


def _ajustar_anchos(hoja, columnas: int, filas: Sequence[Sequence[Any]],
                    titulos: Sequence[str]) -> None:
    """Ancho por columna medido sobre el contenido, acotado por los dos lados."""
    for i in range(columnas):
        ancho = len(str(titulos[i])) if i < len(titulos) else ANCHO_MIN
        for fila in filas:
            if i >= len(fila):
                continue
            v = fila[i]
            if v is None:
                continue
            if isinstance(v, datetime):
                largo = len("2026-01-01 00:00")
            elif isinstance(v, date):
                largo = len("2026-01-01")
            else:
                largo = len(str(v))
            ancho = max(ancho, largo)
        hoja.column_dimensions[get_column_letter(i + 1)].width = (
            max(ANCHO_MIN, min(ANCHO_MAX, ancho + 2)))


# ── Hoja: Cuentas ───────────────────────────────────────────────────────────
def _hoja_cuentas(wb: Workbook, filas: List[Dict[str, Any]]) -> None:
    """Una fila por cuenta, con las columnas del CSV.

    Congelar la cabecera y activar el autofiltro es lo que de verdad hace
    usable la hoja con muchas filas: sin eso, a la fila 40 ya no sabes qué
    columna estás mirando.
    """
    hoja = wb.create_sheet(tabular.nombre_hoja("Cuentas"))
    _fila_cabecera(hoja, tabular.TITULOS)

    matriz: List[List[Any]] = []
    for n, fila in enumerate(filas, start=2):
        valores: List[Any] = []
        for i, col in enumerate(tabular.COLUMNAS, start=1):
            valor = fila.get(col.clave)
            celda = hoja.cell(row=n, column=i)
            _escribir(celda, valor)
            if (col.tipo == "url" and isinstance(valor, str)
                    and valor.startswith(_ESQUEMAS)):
                celda.hyperlink = valor
                celda.style = "Hyperlink"
            valores.append(valor)
        matriz.append(valores)

    hoja.freeze_panes = "A2"
    # El autofiltro necesita un rango; con 0 filas se aplica solo a la cabecera.
    ultima_col = get_column_letter(len(tabular.COLUMNAS))
    hoja.auto_filter.ref = f"A1:{ultima_col}{max(1, len(filas) + 1)}"
    _ajustar_anchos(hoja, len(tabular.COLUMNAS), matriz, tabular.TITULOS)


# ── Hoja: Resumen ───────────────────────────────────────────────────────────
def _alcance(accounts: Sequence[dict]) -> Tuple[List[str], List[str]]:
    """Identificadores del informe, separados en correos y usernames.

    Igual que el PDF: no hay tabla de "qué se escaneó", se deriva de lo que hay.
    El informe declara su propio contenido, no una intención pasada.
    """
    correos, usuarios = set(), set()
    for a in accounts:
        ident = (a.get("identifier") or "").strip()
        if not ident:
            continue
        (correos if "@" in ident else usuarios).add(ident)
    return sorted(correos), sorted(usuarios)


def _distribucion(conteos: Dict[str, int], etiquetar) -> List[Tuple[str, int]]:
    """`{clave: n}` → `[(etiqueta legible, n)]`, de mayor a menor."""
    filas = [(etiquetar(k), int(v or 0)) for k, v in (conteos or {}).items()]
    return sorted(filas, key=lambda p: (-p[1], p[0]))


def _hoja_resumen(wb: Workbook, accounts: Sequence[dict],
                  filas: List[Dict[str, Any]], summary: Dict[str, Any],
                  audit_summary: Dict[str, int], generated_at: float,
                  incidencias: tabular.Incidencias) -> None:
    hoja = wb.create_sheet(tabular.nombre_hoja("Resumen"))
    n = 1

    def linea(texto: Any = None, fuente: Optional[Font] = None,
              valor: Any = None) -> None:
        nonlocal n
        if texto is not None:
            c = hoja.cell(row=n, column=1)
            _escribir(c, texto)
            if fuente:
                c.font = fuente
        if valor is not None:
            _escribir(hoja.cell(row=n, column=2), valor)
        n += 1

    def bloque(titulo: str, pares: Sequence[Tuple[str, int]],
               cabecera: str = "Cuentas") -> None:
        nonlocal n
        linea(titulo, FUENTE_GRUPO)
        _fila_cabecera(hoja, (titulo, cabecera), fila=n)
        # La cabecera del bloque repite el título en A: así cada tabla se
        # entiende sola si alguien copia solo ese trozo.
        n += 1
        for etiqueta, cuenta in pares:
            linea(etiqueta, valor=cuenta)
        linea()

    ts = _t.strftime("%Y-%m-%d %H:%M", _t.gmtime(generated_at))
    linea("Rastrillo · informe de cuentas", FUENTE_TITULO)
    linea(f"Generado el {ts} UTC", FUENTE_NOTA)
    linea()

    correos, usuarios = _alcance(accounts)
    linea("Alcance", FUENTE_GRUPO)
    linea("Cuentas detectadas", valor=int(summary.get("total", 0) or 0))
    linea("Nombres de usuario", valor=len(usuarios))
    linea("Correos", valor=len(correos))
    if usuarios:
        linea("Usuarios", valor=_lista_resumida(usuarios))
    if correos:
        linea("Correos analizados", valor=_lista_resumida(correos))
    linea()

    bloque("Por estado",
           _distribucion(summary.get("by_status") or {},
                         glosario.etiqueta_estado))

    conf: Dict[str, int] = {}
    veri: Dict[str, int] = {}
    for a in accounts:
        conf[a.get("confidence") or ""] = conf.get(a.get("confidence") or "", 0) + 1
        veri[a.get("verifiability") or ""] = veri.get(
            a.get("verifiability") or "", 0) + 1
    bloque("Por confianza",
           _distribucion(conf, glosario.etiqueta_confianza))
    bloque("Por verificabilidad",
           _distribucion(veri, glosario.etiqueta_verificabilidad))

    if audit_summary:
        bloque("Registro de auditoría",
               sorted(((k, int(v)) for k, v in audit_summary.items()),
                      key=lambda p: (-p[1], p[0])),
               cabecera="Veces")
        linea("Total de acciones registradas",
              valor=int(summary.get("audit_total", 0) or 0))
        linea()

    # Notas del informe. Lo que se ha tocado se declara: un recorte que nadie
    # cuenta es una alteración silenciosa del dato.
    linea("Notas", FUENTE_GRUPO)
    linea("Todas las fechas están en UTC.", FUENTE_NOTA)
    linea("«Perfil detectado» es donde se encontró la cuenta; «Cómo darse de "
          "baja», adónde ir para cerrarla. Son cosas distintas.", FUENTE_NOTA)
    linea("«Confianza» dice cuánta evidencia hay de que la cuenta sea tuya; "
          "«Verificabilidad» dice si la respuesta del sitio sirve para "
          "comprobar algo. Son ejes separados.", FUENTE_NOTA)
    if incidencias.recortados:
        linea(f"{incidencias.recortados} valor(es) superaban el máximo de "
              f"{tabular.MAX_CELDA} caracteres por celda y se recortaron; el "
              f"aviso dentro de la celda dice cuántos caracteres faltan.",
              FUENTE_NOTA)
    if incidencias.neutralizados:
        linea(f"{incidencias.neutralizados} valor(es) empezaban por =, +, - o "
              f"@ y se escribieron como texto para que Excel no los ejecute "
              f"como fórmula. El contenido no se ha alterado.", FUENTE_NOTA)

    hoja.column_dimensions["A"].width = 46
    hoja.column_dimensions["B"].width = 30


# ── Hoja: Glosario ──────────────────────────────────────────────────────────
def _hoja_glosario(wb: Workbook) -> None:
    """Las mismas explicaciones que el anexo del PDF.

    No son una copia: se leen de `static/app.js` en tiempo de render, igual que
    el anexo. Si alguien retoca un tooltip del dashboard, el siguiente XLSX lo
    dice igual sin que nadie tenga que acordarse.
    """
    hoja = wb.create_sheet(tabular.nombre_hoja("Glosario"))
    n = 1

    def linea(a: Any = None, b: Any = None, fuente: Optional[Font] = None) -> None:
        nonlocal n
        if a is not None:
            c = hoja.cell(row=n, column=1)
            _escribir(c, a)
            c.font = fuente or Font(bold=True)
        if b is not None:
            c = hoja.cell(row=n, column=2)
            _escribir(c, b)
            c.alignment = Alignment(vertical="top", wrap_text=True)
        n += 1

    linea("Cómo leer este informe", None, FUENTE_TITULO)
    n += 1

    fallo = glosario.error_textos()
    if fallo:
        # Degradación visible: antes que inventar frases, decir que faltan.
        linea("Glosario no disponible", None, FUENTE_GRUPO)
        hoja.cell(row=n, column=1, value=(
            "No se pudieron leer las definiciones desde la interfaz "
            f"({fallo}). Los términos aparecen sin su explicación; "
            "consúltalos en el dashboard.")).font = FUENTE_NOTA
        hoja.column_dimensions["A"].width = 46
        hoja.column_dimensions["B"].width = 96
        return

    linea("Confianza: ¿es mía esta cuenta?", None, FUENTE_GRUPO)
    linea(None, "Mide cuánta evidencia hay de que la cuenta sea tuya. Ninguna "
                "de las señales comprueba nada en el sitio: la decisión sigue "
                "siendo tuya.")
    for nivel in glosario.ORDEN_CONFIANZA:
        linea(glosario.etiqueta_confianza(nivel),
              glosario.explicacion_confianza(nivel))
    n += 1

    linea("Verificabilidad: ¿sirve la respuesta del sitio?", None, FUENTE_GRUPO)
    linea(None, "Eje SEPARADO de la confianza. Habla del SITIO, no de tu "
                "cuenta: que un sitio no sepa distinguir usuarios inexistentes "
                "no dice nada sobre si la cuenta es tuya. Cuando no aparece un "
                "veredicto es que el canario no llegó a mirar ese sitio, que "
                "no es lo mismo que mirarlo y no poder concluir.")
    for code in glosario.MOTIVOS_CANARIO:
        texto = glosario.explicacion_motivo(code)
        if texto:
            linea(glosario.etiqueta_motivo(code) or code, texto)
    n += 1

    linea("Señales de confianza", None, FUENTE_GRUPO)
    linea(None, "Cada cuenta lista las señales que la llevaron a su tramo. "
                "Estas son todas las que Rastrillo sabe registrar.")
    ya = set(glosario.MOTIVOS_CANARIO)
    textos = glosario.cargar_textos().get("REASON_TIP", {})
    for code in sorted(textos):
        if code not in ya:
            linea(glosario.etiqueta_motivo(code) or code, textos[code])
    n += 1

    linea("Estados", None, FUENTE_GRUPO)
    for code, (etiqueta, _color) in glosario.STATUS_META.items():
        linea(etiqueta, code)

    hoja.column_dimensions["A"].width = 46
    hoja.column_dimensions["B"].width = 96


# ── API pública ─────────────────────────────────────────────────────────────
def render_xlsx(accounts: Sequence[dict], summary: Dict[str, Any],
                audit_summary: Dict[str, int], generated_at: float) -> bytes:
    """Construye el libro completo y lo devuelve como bytes.

    Misma firma que `report_pdf.render_pdf` para que `reports.build_report`
    trate los dos formatos igual.
    """
    incidencias = tabular.Incidencias()
    filas = tabular.proyectar(accounts, incidencias)

    wb = Workbook()
    # `Workbook()` nace con una hoja por defecto que no queremos: las tres
    # nuestras se crean con su nombre saneado.
    wb.remove(wb.active)

    _hoja_resumen(wb, accounts, filas, summary, audit_summary, generated_at,
                  incidencias)
    _hoja_cuentas(wb, filas)
    _hoja_glosario(wb)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
